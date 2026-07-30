"""
generar_acum2025.py  (Europa)
-----------------------------
Arma Acumulado_interanual.xlsx (comparativo del anio anterior) a partir del
master europeo data/Ventas_Master_2025.xlsx, SIN tocar report.py.

Diferencias con USA:
  - El master europeo tiene columnas: Local | Grupo | Fecha | Dia | Venta Bruta (EUR).
  - La venta viene en BRUTO: la paso a NETO dividiendo por (1 + IVA) = 1.10, para
    comparar contra los cierres, que report.py tambien muestra en neto.
  - Solo 7 locales tienen historial 2025. Madrid y Alicante abrieron en 2026:
    no se escriben. report.py los deja en 0 y los muestra como "—" (via
    SIN_HISTORICO_2025). El unico lugar donde vive "quien no tiene 2025" es
    report.py; aca solo escribimos lo que el master tiene.

Escribe el MISMO formato que espera report.py: fila 0 = titulo con rango,
fila 1 = header, filas 2+ = local y Net Sales en col C. Los nombres de local
del master ya son los canonicos ("Barcelona 1", "Malaga 1", ...).

Falla RUIDOSA: si el master no tiene datos para ese mes, corta con error.
"""
import re
import sys
from pathlib import Path
from datetime import date, datetime
from collections import defaultdict

import openpyxl

BASE = Path(__file__).parent
VENTAS = BASE / "Ventas_ayer.xlsx"
MASTER = BASE / "data" / "Ventas_Master_2025.xlsx"
SALIDA = BASE / "Acumulado_interanual.xlsx"
IVA = 0.10  # neto = bruto / (1 + IVA)


def _to_float(v):
    if v is None:
        return 0.0
    return float(str(v).replace("$", "").strip() or 0)


def _as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()


def dia_de_corte():
    """Lee la fecha fin del titulo de Ventas_ayer.xlsx."""
    wb = openpyxl.load_workbook(VENTAS, data_only=True)
    ws = wb.active
    title = str(next(ws.iter_rows(values_only=True))[0] or "")
    m = re.search(r"(\d{4}-\d{2}-\d{2})\s*/\s*(\d{4}-\d{2}-\d{2})", title)
    if not m:
        raise ValueError(f"No pude leer la fecha de Ventas_ayer.xlsx: {title!r}")
    return datetime.strptime(m.group(2), "%Y-%m-%d").date()


def acumular_2025(corte):
    """Suma del master la venta NETA por local, del 1ro del mes del anio
    anterior hasta el mismo dia de corte."""
    ini = date(corte.year - 1, corte.month, 1)
    fin = date(corte.year - 1, corte.month, corte.day)

    wb = openpyxl.load_workbook(MASTER, data_only=True)
    ws = wb["Ventas 2025"]
    header = [str(c) for c in next(ws.iter_rows(values_only=True))]
    col_local = header.index("Local")
    col_fecha = header.index("Fecha")
    col_venta = header.index("Venta Bruta (EUR)")

    acum = defaultdict(float)
    hubo_datos = False
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[col_local] is None:
            continue
        f = _as_date(r[col_fecha])
        if ini <= f <= fin:
            neto = _to_float(r[col_venta]) / (1 + IVA)
            acum[str(r[col_local]).strip()] += neto
            hubo_datos = True

    if not hubo_datos:
        raise SystemExit(
            f"[ERROR] El master 2025 no tiene datos para {ini}..{fin}. "
            f"Cargale ese mes al master antes de correr."
        )
    return ini, fin, acum


def escribir(ini, fin, acum):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = f"Ventas {ini.isoformat()} / {fin.isoformat()}"
    ws.append(["Sucursal", "", "Net Sales", "", "", "Bill Count"])
    total = 0.0
    for local in sorted(acum):
        v = round(acum[local], 2)
        total += v
        ws.append([local, "", v, "", "", 0])
    wb.save(SALIDA)
    return total


if __name__ == "__main__":
    corte = dia_de_corte()
    ini, fin, acum = acumular_2025(corte)
    tot = escribir(ini, fin, acum)
    print(f"OK - Acumulado 2025 (neto) {ini}..{fin} | {len(acum)} locales | total {tot:,.2f} EUR")
