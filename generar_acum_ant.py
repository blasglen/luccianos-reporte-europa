"""
generar_acum_ant.py  (Europa)
-----------------------------
Comparativo del AÑO ANTERIOR para CUALQUIER rango de fechas, leyendo el master
diario data/Ventas_Master_2025.xlsx (hoja "Ventas 2025":
Local | Fecha | Venta Bruta (EUR) | Tickets).

Lo usan el SEMANAL (rango lunes-domingo) y el CIERRE (rango del mes).
Venta en NETO (bruto ÷ 1.10). Locales sin datos en el rango (Madrid/Alicante en
2025) simplemente no aparecen; report.py los toma como "sin comparable".

Interfaz (la que esperan semanal y cierre):
  acumular_rango(desde, hasta) -> (acum_neto:dict, tickets:dict, faltan:list[date])
  escribir_excel(desde, hasta, acum, salida, tickets)  -> Net Sales (C) + Bill Count (F)
  espejo(d) -> mismo dia/mes, año-1
"""
import argparse
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from collections import defaultdict

import openpyxl

BASE = Path(__file__).parent
MASTER = BASE / "data" / "Ventas_Master_2025.xlsx"
IVA = 0.10

# Los 9 locales canonicos (para inicializar en cero de forma estable)
BRANCHES = ["Barcelona 1", "Barcelona 2", "Madrid", "Roma",
            "Málaga 1", "Málaga 3", "Valencia", "Alicante", "Granada"]


def _as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()


def espejo(d):
    """Mismo dia y mes, año anterior."""
    return date(d.year - 1, d.month, d.day)


def _leer_master():
    wb = openpyxl.load_workbook(MASTER, data_only=True)
    ws = wb["Ventas 2025"]
    h = [str(c) for c in next(ws.iter_rows(values_only=True))]
    cl, cf, cv = h.index("Local"), h.index("Fecha"), h.index("Venta Bruta (EUR)")
    ct = h.index("Tickets") if "Tickets" in h else None
    por_dia = defaultdict(dict)   # fecha -> {local: (bruto, tickets)}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[cl] is None:
            continue
        f = _as_date(r[cf])
        tk = int(r[ct] or 0) if ct is not None else 0
        por_dia[f][str(r[cl]).strip()] = (float(r[cv] or 0), tk)
    return por_dia


def acumular_rango(desde, hasta):
    """Suma NETA y tickets por local entre [desde, hasta] inclusive.
    Devuelve (acum, tickets, faltan). 'faltan' = dias del rango sin ninguna fila
    en el master (el caller decide si aborta)."""
    por_dia = _leer_master()
    acum = {b: 0.0 for b in BRANCHES}
    tickets = {b: 0 for b in BRANCHES}
    faltan = []
    d = desde
    while d <= hasta:
        if d not in por_dia:
            faltan.append(d)
        else:
            for loc, (bruto, tk) in por_dia[d].items():
                acum.setdefault(loc, 0.0)
                tickets.setdefault(loc, 0)
                acum[loc] += round(bruto / (1 + IVA), 2)
                tickets[loc] += tk
        d += timedelta(days=1)
    acum = {k: round(v, 2) for k, v in acum.items()}
    return acum, tickets, faltan


def escribir_excel(desde, hasta, acum, salida, tickets=None):
    tickets = tickets or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = f"Ventas {desde.isoformat()} / {hasta.isoformat()}"
    ws.append(["Sucursal", "", "Net Sales", "", "", "Bill Count"])
    for loc in sorted(acum):
        ws.append([loc, "", round(acum[loc], 2), "", "", int(tickets.get(loc, 0))])
    wb.save(salida)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--desde", required=True)
    ap.add_argument("--hasta", required=True)
    ap.add_argument("--salida", required=True)
    a = ap.parse_args()
    desde = datetime.strptime(a.desde, "%Y-%m-%d").date()
    hasta = datetime.strptime(a.hasta, "%Y-%m-%d").date()
    acum, tickets, faltan = acumular_rango(desde, hasta)
    if faltan:
        print(f"[AVISO] faltan {len(faltan)} dias en el master: "
              f"{', '.join(x.isoformat() for x in faltan[:8])}"
              f"{'...' if len(faltan) > 8 else ''}")
    escribir_excel(desde, hasta, acum, a.salida, tickets)
    print(f"OK - comparativo {desde}..{hasta} (neto) | {len([k for k in acum if acum[k]])} locales con datos | "
          f"total {sum(acum.values()):,.2f} EUR")
